import sys
from unittest.util import strclass
import pandas as pd
import numpy as np
from PyQt6.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QWidget, QFileDialog, QPushButton, QLabel, QListWidget, QAbstractItemView
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class FileComparator(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Spreadsheet Comparator")
        self.setGeometry(300, 100, 500, 300)
        
        self.init_ui()

    def init_ui(self):
        # Set up layout and widgets
        self.layout = QVBoxLayout()
        
        self.info_label = QLabel("Select two files to compare:")
        self.layout.addWidget(self.info_label)
        
        self.select_button_1 = QPushButton("Select First File")
        self.select_button_1.clicked.connect(self.select_file_1)
        self.layout.addWidget(self.select_button_1)
        
        self.select_button_2 = QPushButton("Select Second File")
        self.select_button_2.clicked.connect(self.select_file_2)
        self.layout.addWidget(self.select_button_2)

        # Create a label for the multi-select combo box
        self.combo_box_label = QLabel("Select Common Column Headers")
        self.layout.addWidget(self.combo_box_label)

        # Create a multi-select list widget
        self.multi_select_combo_box = QListWidget()
        self.multi_select_combo_box.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        self.layout.addWidget(self.multi_select_combo_box)

        # Create a button to print selected items
        self.print_button = QPushButton("Common Columns Selected")
        self.print_button.clicked.connect(self.print_selected_columns)
        self.layout.addWidget(self.print_button)
        
        self.compare_button = QPushButton("Compare Files")
        self.compare_button.clicked.connect(self.compare_files)
        self.compare_button.setEnabled(False)
        self.layout.addWidget(self.compare_button)
        
        self.result_label = QLabel("")
        self.layout.addWidget(self.result_label)
        
        container = QWidget()
        container.setLayout(self.layout)
        self.setCentralWidget(container)
        
        self.file1 = None
        self.file2 = None

    def select_file_1(self):
        # Open file dialog for selecting the first file
        self.file1, _ = QFileDialog.getOpenFileName(self, "Select First File", "", "All Files (*);;Excel Files (*.xlsx *.xls);;CSV Files (*.csv)")
        if self.file1:
            self.select_button_1.setText(f"Selected: {self.file1}")
            self.check_files_ready()
            self.load_columns()

    def select_file_2(self):
        # Open file dialog for selecting the second file
        self.file2, _ = QFileDialog.getOpenFileName(self, "Select Second File", "", "All Files (*);;Excel Files (*.xlsx *.xls);;CSV Files (*.csv)")
        if self.file2:
            self.select_button_2.setText(f"Selected: {self.file2}")
            self.check_files_ready()

    def check_files_ready(self):
        # Enable the compare button only if both files are selected
        if self.file1 and self.file2:
            self.compare_button.setEnabled(True)

    def load_columns(self):
        try:
            # Read the file to extract column headers
            if self.file1.endswith('.csv'):
                data = pd.read_csv(self.file1)
            else:
                data = pd.read_excel(self.file1)

            # Get the column headers
            column_headers = list(data.columns)

            # Populate the multi-select list widget
            self.multi_select_combo_box.clear()
            self.multi_select_combo_box.addItems(column_headers)

        except Exception as e:
            print(f"Error reading file: {e}")

    def print_selected_columns(self):
        # Get selected items
        selected_items = [item.text() for item in self.multi_select_combo_box.selectedItems()]
        print("Selected Columns:", selected_items)

    def gen_col_type(self,num_keys):
        return {i: 'str' for i in range(num_keys)}

    def compare_files(self):
        try:
            # Check file types and read accordingly

            if self.file1.endswith('.csv') and self.file2.endswith('.csv'):
                df1 = pd.read_csv(self.file1, dtype = str)
                df2 = pd.read_csv(self.file2, dtype = str)
                self.compare_csv(df1, df2)
            else:
                # Make sure all columns are read as type object
                pdf = pd.read_excel(self.file1)
                ncol1 = len(pdf.axes[1])
                clmtype = self.gen_col_type(ncol1)
                df1 = pd.read_excel(self.file1,dtype= clmtype)
                df1 = df1.fillna('')
                #print("DF1")
                #print(df1.info())
                pdf = pd.read_excel(self.file2)
                ncol2 = len(pdf.axes[1])
                clmtype = self.gen_col_type(ncol2)
                df2 = pd.read_excel(self.file2,dtype= clmtype)
                df2 = df2.fillna('')
                #print("DF2")
                #print(df2.info())
                self.compare_excel(df1, df2)
        except Exception as e:
            self.result_label.setText(f"Error: {e}")

    def strip(self,x):
        """Function to use with applymap to strip whitespaces in a dataframe."""
        return x.strip() if isinstance(x, str) else x


    def report_diff(self,x):
        """Function to use with groupby.apply to highlight value changes."""
        return x[0] if x[0] == x[1] or pd.isna(x).all() else f'{x[0]} ---> {x[1]}'

    def highlight_diff(self,x):
        """Function to use with style.applymap to highlight value changes."""
        if isinstance(x,str):
            if "--->" in x:
                return 'background-color: orange'
                #return 'background-color: %s' % 'orange'
        return ''

    def diff_pd(self,df1,df2):
        idx_col = [item.text() for item in self.multi_select_combo_box.selectedItems()]
        # setting the column name as index for fast operations
        df1 = df1.set_index(idx_col)
        df2 = df2.set_index(idx_col)

        # get the added and removed rows
        file1_keys = df1.index
        #file1_keys[:] = [item if item != nan else '' for item in file1_keys]
        file2_keys = df2.index
        if isinstance(file1_keys, pd.MultiIndex):
            removed_keys = file1_keys.difference(file2_keys)
            added_keys = file2_keys.difference(file1_keys)
        else:
            removed_keys = np.setdiff1d(file1_keys, file2_keys)
            added_keys = np.setdiff1d(file2_keys, file1_keys)
        out_data = {
            'removed row': df1.loc[removed_keys],
            'added row': df2.loc[added_keys]
        }
        # focusing on common data of both dataframes
        common_keys = np.intersect1d(file1_keys, file2_keys, assume_unique=True)
        common_columns = np.intersect1d(
            df1.columns, df2.columns, assume_unique=True
        )
        file2_common = df2.loc[common_keys, common_columns].applymap(self.strip)
        file1_common = df1.loc[common_keys, common_columns].applymap(self.strip)
        # get the changed rows keys by dropping identical rows
        # (indexes are ignored, so we'll reset them)
        common_data = pd.concat(
            [file1_common.reset_index(), file2_common.reset_index()], sort=True
        )
        changed_keys = common_data.drop_duplicates(keep=False)[idx_col]
        if isinstance(changed_keys, pd.Series):
            changed_keys = changed_keys.unique()
        else:
            changed_keys = changed_keys.drop_duplicates().set_index(idx_col).index
        # combining the changed rows via multi level columns
        df_all_changes = pd.concat(
            [file1_common.loc[changed_keys], file2_common.loc[changed_keys]],
            axis='columns',
            keys=['file1', 'file2']
        ).swaplevel(axis='columns')
        # using report_diff to merge the changes in a single cell with "-->"
        df_changed = df_all_changes.groupby(level=0, axis=1).apply(
            lambda frame: frame.apply(self.report_diff, axis=1))

        if len(df_changed.index) > 0:
            styleddf = (df_changed.style.map(self.highlight_diff))
        #out_data['changed row'] = df_changed
            out_data['changed row'] = styleddf

        # get the added and removed columns
        file1_col = list(df1.columns)
        file2_col = list(df2.columns)
        removed_col = np.setdiff1d(file1_col, file2_col)
        added_col = np.setdiff1d(file2_col, file1_col)

        if(removed_col.size > 0):
            file1_diff = df1.loc[common_keys, removed_col].applymap(self.strip)
            out_data['removed col'] = file1_diff

        if(added_col.size > 0):
            file2_diff = df2.loc[common_keys, added_col].applymap(self.strip)
            out_data['added col'] = file2_diff

        return out_data

    def compare_excel(self, df1, df2):
        # Compare two Excel files even with different structures (rows/columns)
        diff = self.diff_pd(df1, df2)
        with pd.ExcelWriter('Compared_File.xlsx') as writer:
            for sname, data in diff.items():
                data.to_excel(writer, engine='openpyxl',sheet_name=sname)
        
        #if diff.empty:
            #self.result_label.setText("The files are identical.")
            #return
        
        # Highlight differences in Excel
        #self.highlight_excel_differences(df1, df2, diff_df)
        self.result_label.setText("Differences found and highlighted in 'Compared_File.xlsx'.")

    def merge_excel(self, df1, df2):
        # Compare two Excel files even with different structures (rows/columns)
        diff_df = pd.merge(df1, df2, how='outer', indicator=True)
        
        if diff_df.empty:
            self.result_label.setText("The files are identical.")
            return
        
        # Highlight differences in Excel
        self.highlight_excel_differences(df1, df2, diff_df)
        self.result_label.setText("Differences found and highlighted in 'Compared_File.xlsx'.")

    def compare_csv(self, df1, df2):
        # Compare two CSV files even with different structures (rows/columns)
        df1 = df1.fillna('')
        df1 = df1.applymap(str)
        df2 = df2.fillna('')
        df2 = df2.applymap(str)
        # Compare the two dataframes, marking any differences
        #diff_df = pd.concat([df1, df2]).drop_duplicates(keep=False)
        diff = self.diff_pd(df1, df2)
        with pd.ExcelWriter('Compared_File.xlsx') as writer:
            for sname, data in diff.items():
                data.to_excel(writer, engine='openpyxl',sheet_name=sname)        
        
        #if diff_df.empty:
            #self.result_label.setText("The files are identical.")
            #return
        
        # Save differences to CSV file
        #self.highlight_csv_differences(diff_df)
        self.result_label.setText("Differences found and saved in 'Compared_File.xlsx'.")

    def highlight_excel_differences(self, df1, df2, diff_df):
        # Load the original Excel file
        wb = load_workbook(self.file1)
        ws = wb.active
        
        # Define a yellow fill for differences
        highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Iterate over the differences and apply highlights
        for row in range(len(df1)):
            for col in range(len(df1.columns)):
                if diff_df.iloc[row, col] != diff_df.iloc[row, col]:  # This checks for NaNs in the diff
                    ws.cell(row=row+2, column=col+1).fill = highlight  # +2 to account for header row
        
        # Save the highlighted file
        wb.save('Compared_File.xlsx')

    def highlight_csv_differences(self, diff_df):
        # Save differences to CSV with marking
        diff_file = 'Compared_File.csv'
        diff_df.to_csv(diff_file, index=False)

        self.result_label.setText(f"Differences found and saved in {diff_file}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = FileComparator()
    window.show()
    sys.exit(app.exec())